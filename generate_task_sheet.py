import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

tasks = [
    # (#, Page/Section, Task, Details, Type)
    (1, "Login & Access", "Create login screen", "Email + password login screen. Currently the app has no login - users land directly on the dashboard.", "New"),
    (2, "Login & Access", "Role-based company logo", "After login, show the logo of the user's own company. Niranjan (Group HR) sees the ABG group logo. A user from Hindalco sees only the Hindalco logo and only Hindalco data.", "New"),
    (3, "Login & Access", "Role-based data scoping", "When a company-level user logs in, all dashboard data (clusters, heatmap, BUs) must be filtered to only their company. They see their BUs, not the full group view.", "New"),
    (4, "Login & Access", "Logout button", "Add a visible logout option in the sidebar or header. Move existing elements slightly to accommodate.", "New"),

    (5, "Overview Page", "Key takeaway - collapse by default", "Show the yellow Key Takeaway box first (after AI executive summary). Hide the three detail bullets by default. User clicks \"More\" / expand to reveal them. Clicking again collapses back.", "Enhance"),
    (6, "Overview Page", "BU cluster cards - collapse Top BUs", "In each of the four cluster cards (Thriving, At Risk, Polarised, Critical), hide the \"Top BUs\" list by default. Show only a dropdown/expand arrow. Clicking it expands to show the list.", "Enhance"),
    (7, "Overview Page", "AI Recommended Focus Areas - remove trend line", "The small trend line inside each focus area card (Critical Watchlist, Emerging Risks, Bright Spots) should be removed. Replace with something grounded in existing data - e.g. a score badge or category breakdown. No placeholder time-series charts.", "Fix"),
    (8, "Overview Page", "Make filters functional", "The Filters button (Cluster, Driver) currently doesn't work. Make it active - applying a cluster or driver filter should update the dashboard view accordingly.", "Fix"),
    (9, "Overview Page", "Add company + BU filter", "Extend filters to include: (1) filter by company - for Group HR users like Niranjan who want to see e.g. only mining companies. (2) Filter by BU within a company. Company-level users only need BU filter (no company selector needed for them).", "New"),

    (10, "Business Overview Page", "Remove trend lines from company cards", "The small green trend lines shown in each company card are not meaningful (no time-series data). Remove them entirely.", "Remove"),
    (11, "Business Overview Page", "Add sort options", "Add sort controls with two options: (1) sort by engagement score (high -> low or low -> high) and (2) sort by number of responses. Both should support ascending and descending order. User picks one at a time.", "New"),

    (12, "Business Detail Page", "Add value labels to radar / hexagon chart", "The radar chart categories (Development, Leadership, Performance Culture, etc.) currently have no visible values. Add hover tooltips showing the exact score (e.g. 4.5) for each dimension.", "Enhance"),

    (13, "Insights Studio (AI Chat)", "Remove the word \"struggling\"", "The sample question \"Why is ABG struggling?\" uses negative language. Replace with neutral phrasing - e.g. \"What is driving engagement at ABG?\" or similar.", "Fix"),
    (14, "Insights Studio (AI Chat)", "Add two dimension filters to chat", "Allow the user to add context to their AI query via two filters: (1) a dimension filter (e.g. psychological safety, gender) and (2) a company filter. Example: \"Analyse psychological safety by gender for Fashion Retail.\" Implement if feasible.", "New"),

    (15, "Trends Over Time Page", "Replace Jan/Feb/Mar with wave labels", "X-axis labels currently show months. Replace with survey wave names: \"Vibes 2024\", \"Vibes 2025\", etc. Use dummy data - no actual historical data exists yet.", "Fix"),
    (16, "Trends Over Time Page", "Add drill-down into trend chart", "Allow clicking on the overall trend line to drill down - e.g. see the same trend broken down by topic (Leadership, Reward, etc.) or by company. Can be static/dummy for now.", "Enhance"),

    (17, "Outliers & Alerts Page", "Add threshold selector", "Add a control to let the user define what counts as \"top performing\" or \"critical\" - e.g. a dropdown or input with a default value. Used to classify BUs in the outliers view.", "New"),

    (18, "Employee Voice Page", "Move sentiment cards above bar chart", "The three sentiment cards (Positive, Mixed, Negative) should appear above the bar graph, not below it.", "Enhance"),
    (19, "Employee Voice Page", "Expand sentiment cards to full width", "The three sentiment cards currently don't use the full page width. Stretch them to use all available horizontal space.", "Enhance"),
    (20, "Employee Voice Page", "Add filters (cohort, company, BU)", "Add three filters: (1) cohort - tenure, job band, gender etc., (2) company, and (3) BU within a company. These allow the user to drill into sentiment for a specific group or location.", "New"),

    (21, "Settings Page", "Minimum sample size - placeholder screen", "Add a \"Minimum sample size\" setting in the Settings page. No functional logic needed yet - just a placeholder UI to show the concept. A note can explain that real data segmentation will be added once clean data is available.", "Placeholder"),

    (22, "Sidebar / Navigation", "Add three new Explore sub-pages", "Aditya will share wireframes for three new pages to be added under the Explore section: (1) Hypothesis Testing, (2) Statistical Analysis, (3) Sentiment Analysis. Frontend to be built once wireframes are received.", "New - Awaiting Wireframes"),

    (23, "Global / Polish", "Standardise font sizes", "Ensure font sizes are visually consistent across all pages. The \"All Insights\" label and other small labels in the right panel appear too small compared to surrounding text. Do a pass across all pages.", "Fix"),
]

df = pd.DataFrame(tasks, columns=["#", "Page / Section", "Task", "Details", "Type"])
df["Assignee"] = ""
df["Priority"] = ""
df["Status"] = "To Do"
df["Due Date"] = ""
df["Notes"] = ""

output_path = "Dashboard_Task_Management.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Tasks", index=False)

    ws = writer.sheets["Tasks"]

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Column widths
    widths = {
        "A": 5,   # #
        "B": 22,  # Page/Section
        "C": 32,  # Task
        "D": 70,  # Details
        "E": 18,  # Type
        "F": 16,  # Assignee
        "G": 12,  # Priority
        "H": 14,  # Status
        "I": 14,  # Due Date
        "J": 30,  # Notes
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Wrap text + top align for Task/Details columns, row heights
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row[0].row].height = 60

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

print(f"Created {output_path}")
